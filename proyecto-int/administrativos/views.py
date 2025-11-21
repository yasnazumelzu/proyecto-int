from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q, Avg
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime, timedelta
from pacientes.models import Paciente, RecienNacido
from .models import Admision, Pago, AltaEgreso
from .forms import AdmisionForm, PagoForm, AltaEgresoForm
import io
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

@login_required
def dashboard_admin(request):
    """Dashboard principal del administrativo"""
    pacientes = Paciente.objects.all()
    
    # Estadísticas de admisiones
    total_admisiones = Admision.objects.count()
    admisiones_pendientes = Admision.objects.filter(estado='pendiente').count()
    admisiones_hoy = Admision.objects.filter(fecha_admision__date=timezone.now().date()).count()
    
    # Estadísticas de pagos
    total_pagos = Pago.objects.count()
    pagos_pendientes = Pago.objects.filter(estado='pendiente').count()
    monto_total_pendiente = Pago.objects.filter(estado__in=['pendiente', 'parcial']).aggregate(
        total=Sum('monto_pendiente')
    )['total'] or 0
    monto_total_pagado_hoy = Pago.objects.filter(
        fecha_pago__date=timezone.now().date(),
        estado='pagado'
    ).aggregate(total=Sum('monto_pagado'))['total'] or 0
    
    context = {
        'pacientes': pacientes,
        'total_admisiones': total_admisiones,
        'admisiones_pendientes': admisiones_pendientes,
        'admisiones_hoy': admisiones_hoy,
        'total_pagos': total_pagos,
        'pagos_pendientes': pagos_pendientes,
        'monto_total_pendiente': monto_total_pendiente,
        'monto_total_pagado_hoy': monto_total_pagado_hoy,
    }
    return render(request, 'administrativo/dashboard.html', context)

@login_required
def lista_admisiones(request):
    """Lista todas las admisiones"""
    admisiones = Admision.objects.all().select_related('paciente', 'usuario_registro')
    
    # Filtros
    estado_filter = request.GET.get('estado', '')
    tipo_filter = request.GET.get('tipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if estado_filter:
        admisiones = admisiones.filter(estado=estado_filter)
    if tipo_filter:
        admisiones = admisiones.filter(tipo_admision=tipo_filter)
    if fecha_desde:
        admisiones = admisiones.filter(fecha_admision__date__gte=fecha_desde)
    if fecha_hasta:
        admisiones = admisiones.filter(fecha_admision__date__lte=fecha_hasta)
    
    context = {
        'admisiones': admisiones,
        'estado_filter': estado_filter,
        'tipo_filter': tipo_filter,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    return render(request, 'administrativo/lista_admisiones.html', context)

@login_required
def detalle_admision(request, id):
    """Detalle de una admisión específica"""
    admision = get_object_or_404(Admision, id=id)
    pagos = Pago.objects.filter(admision=admision)
    
    context = {
        'admision': admision,
        'pagos': pagos,
    }
    return render(request, 'administrativo/detalle_admision.html', context)

@login_required
def nueva_admision(request):
    """Crear una nueva admisión"""
    if request.method == 'POST':
        form = AdmisionForm(request.POST)
        if form.is_valid():
            admision = form.save(commit=False)
            admision.usuario_registro = request.user
            admision.save()
            return redirect('detalle_admision', id=admision.id)
    else:
        form = AdmisionForm()
    
    return render(request, 'administrativo/nueva_admision.html', {'form': form})

@login_required
def lista_pagos(request):
    """Lista todos los pagos"""
    pagos = Pago.objects.all().select_related('paciente', 'admision', 'usuario_registro')
    
    # Filtros
    estado_filter = request.GET.get('estado', '')
    metodo_filter = request.GET.get('metodo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if estado_filter:
        pagos = pagos.filter(estado=estado_filter)
    if metodo_filter:
        pagos = pagos.filter(metodo_pago=metodo_filter)
    if fecha_desde:
        pagos = pagos.filter(fecha_pago__date__gte=fecha_desde)
    if fecha_hasta:
        pagos = pagos.filter(fecha_pago__date__lte=fecha_hasta)
    
    # Estadísticas
    total_pagado = pagos.filter(estado='pagado').aggregate(
        total=Sum('monto_pagado')
    )['total'] or 0
    total_pendiente = pagos.filter(estado__in=['pendiente', 'parcial']).aggregate(
        total=Sum('monto_pendiente')
    )['total'] or 0
    
    context = {
        'pagos': pagos,
        'estado_filter': estado_filter,
        'metodo_filter': metodo_filter,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'total_pagado': total_pagado,
        'total_pendiente': total_pendiente,
    }
    return render(request, 'administrativo/lista_pagos.html', context)

@login_required
def nuevo_pago(request, admision_id=None):
    """Crear un nuevo pago"""
    admision = None
    if admision_id:
        admision = get_object_or_404(Admision, id=admision_id)
    
    if request.method == 'POST':
        form = PagoForm(request.POST)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.usuario_registro = request.user
            if admision:
                pago.admision = admision
            pago.save()
            return redirect('lista_pagos')
    else:
        initial = {}
        if admision:
            initial['admision'] = admision
            initial['paciente'] = admision.paciente
        form = PagoForm(initial=initial)
    
    context = {
        'form': form,
        'admision': admision,
    }
    return render(request, 'administrativo/nuevo_pago.html', context)

@login_required
def reportes_generales(request):
    """Vista para mostrar reportes estadísticos generales"""
    # Filtros de fecha
    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')
    
    # Si no hay fechas, usar el mes actual
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
        except:
            fecha_desde = timezone.now().replace(day=1).date()
    else:
        fecha_desde = timezone.now().replace(day=1).date()
    
    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
        except:
            fecha_hasta = timezone.now().date()
    else:
        fecha_hasta = timezone.now().date()
    
    # Filtrar recién nacidos por fecha
    recien_nacidos = RecienNacido.objects.filter(
        fecha_nacimiento__range=[fecha_desde, fecha_hasta]
    )
    
    # SECCIÓN D.1: INFORMACIÓN GENERAL DE RECIÉN NACIDOS VIVOS
    total_nacidos_vivos = recien_nacidos.count()
    
    # Peso al nacer (en gramos)
    peso_menor_500 = recien_nacidos.filter(peso__lt=500).count()
    peso_500_999 = recien_nacidos.filter(peso__gte=500, peso__lt=1000).count()
    peso_1000_1499 = recien_nacidos.filter(peso__gte=1000, peso__lt=1500).count()
    peso_1500_1999 = recien_nacidos.filter(peso__gte=1500, peso__lt=2000).count()
    peso_2000_2499 = recien_nacidos.filter(peso__gte=2000, peso__lt=2500).count()
    peso_2500_2999 = recien_nacidos.filter(peso__gte=2500, peso__lt=3000).count()
    peso_3000_3999 = recien_nacidos.filter(peso__gte=3000, peso__lt=4000).count()
    peso_4000_mas = recien_nacidos.filter(peso__gte=4000).count()
    
    # Anomalía congénita
    anomalia_congenita = recien_nacidos.filter(malformacion_congenita=True).count()
    
    # SECCIÓN D.2: ATENCIÓN INMEDIATA DEL RECIÉN NACIDO
    # Profilaxis
    profilaxis_hepatitis_b = recien_nacidos.filter(vacuna_hepatitis_b=True).count()
    profilaxis_ocular = recien_nacidos.filter(profilaxis_ocular=True).count()
    
    # Tipo de parto
    parto_vaginal = recien_nacidos.filter(tipo_parto='normal').count()
    parto_instrumental = recien_nacidos.filter(tipo_parto='instrumental').count()
    cesarea = recien_nacidos.filter(tipo_parto='cesarea').count()
    
    # APGAR
    apgar_1min_menor_igual_3 = recien_nacidos.filter(apgar_1min__lte=3).count()
    apgar_5min_menor_igual_5 = recien_nacidos.filter(apgar_5min__lte=5).count()
    reanimacion_basica_count = recien_nacidos.filter(reanimacion_basica=True).count()
    reanimacion_avanzada_count = recien_nacidos.filter(reanimacion_avanzada=True).count()
    
    # CARACTERÍSTICAS DEL PARTO
    total_partos = recien_nacidos.count()
    partos_vaginal = recien_nacidos.filter(tipo_parto='normal').count()
    partos_instrumental = recien_nacidos.filter(tipo_parto='instrumental').count()
    cesarea_electiva = recien_nacidos.filter(tipo_parto='cesarea').count()  # Simplificado
    cesarea_urgencia = 0  # Necesitaríamos un campo adicional para distinguir
    
    # Partos según edad de la madre
    pacientes_madres = Paciente.objects.filter(
        recien_nacidos__fecha_nacimiento__range=[fecha_desde, fecha_hasta]
    ).distinct()
    
    partos_menor_15 = pacientes_madres.filter(edad__lt=15).count()
    partos_15_19 = pacientes_madres.filter(edad__gte=15, edad__lt=20).count()
    partos_20_34 = pacientes_madres.filter(edad__gte=20, edad__lt=35).count()
    partos_35_mas = pacientes_madres.filter(edad__gte=35).count()
    
    # Partos prematuros (considerar sobre 22 semanas)
    partos_22_24_semanas = recien_nacidos.filter(
        edad_gestacional_semanas__gte=22, edad_gestacional_semanas__lt=24
    ).count()
    partos_24_28_semanas = recien_nacidos.filter(
        edad_gestacional_semanas__gte=24, edad_gestacional_semanas__lt=28
    ).count()
    partos_28_32_semanas = recien_nacidos.filter(
        edad_gestacional_semanas__gte=28, edad_gestacional_semanas__lt=32
    ).count()
    partos_32_37_semanas = recien_nacidos.filter(
        edad_gestacional_semanas__gte=32, edad_gestacional_semanas__lt=37
    ).count()
    
    # Lactancia materna en las primeras 60 minutos
    lactancia_60min = recien_nacidos.filter(
        lactancia_antes_60min=True,
        peso__gte=2500
    ).count()
    
    # Pueblos originarios y migrantes
    pueblos_originarios = pacientes_madres.filter(
        pueblo_originario__isnull=False
    ).exclude(pueblo_originario='').count()
    migrantes = pacientes_madres.filter(es_migrante=True).count()
    
    # Profilaxis ocular para gonorrea
    profilaxis_ocular_total = recien_nacidos.filter(profilaxis_ocular=True).count()
    profilaxis_ocular_pueblos = recien_nacidos.filter(
        profilaxis_ocular=True,
        paciente_madre__pueblo_originario__isnull=False
    ).exclude(paciente_madre__pueblo_originario='').count()
    profilaxis_ocular_migrantes = recien_nacidos.filter(
        profilaxis_ocular=True,
        paciente_madre__es_migrante=True
    ).count()
    
    # Profilaxis Hepatitis B (hijos de madre Hepatitis B positiva)
    madres_hepatitis_b = pacientes_madres.filter(hepatitis_b_resultado__icontains='positivo')
    rn_hepatitis_b = recien_nacidos.filter(paciente_madre__in=madres_hepatitis_b)
    rn_hepatitis_b_con_profilaxis = rn_hepatitis_b.filter(
        vacuna_hepatitis_b=True,
        paciente_madre__hepatitis_b_profilaxis_completa=True
    ).count()
    rn_hepatitis_b_sin_profilaxis = rn_hepatitis_b.count() - rn_hepatitis_b_con_profilaxis
    
    context = {
        'fecha_desde': fecha_desde.strftime('%Y-%m-%d'),
        'fecha_hasta': fecha_hasta.strftime('%Y-%m-%d'),
        # D.1
        'total_nacidos_vivos': total_nacidos_vivos,
        'peso_menor_500': peso_menor_500,
        'peso_500_999': peso_500_999,
        'peso_1000_1499': peso_1000_1499,
        'peso_1500_1999': peso_1500_1999,
        'peso_2000_2499': peso_2000_2499,
        'peso_2500_2999': peso_2500_2999,
        'peso_3000_3999': peso_3000_3999,
        'peso_4000_mas': peso_4000_mas,
        'anomalia_congenita': anomalia_congenita,
        # D.2
        'profilaxis_hepatitis_b': profilaxis_hepatitis_b,
        'profilaxis_ocular': profilaxis_ocular,
        'parto_vaginal': parto_vaginal,
        'parto_instrumental': parto_instrumental,
        'cesarea': cesarea,
        'apgar_1min_menor_igual_3': apgar_1min_menor_igual_3,
        'apgar_5min_menor_igual_5': apgar_5min_menor_igual_5,
        'reanimacion_basica_count': reanimacion_basica_count,
        'reanimacion_avanzada_count': reanimacion_avanzada_count,
        # Características del parto
        'total_partos': total_partos,
        'partos_vaginal': partos_vaginal,
        'partos_instrumental': partos_instrumental,
        'cesarea_electiva': cesarea_electiva,
        'cesarea_urgencia': cesarea_urgencia,
        'partos_menor_15': partos_menor_15,
        'partos_15_19': partos_15_19,
        'partos_20_34': partos_20_34,
        'partos_35_mas': partos_35_mas,
        'partos_22_24_semanas': partos_22_24_semanas,
        'partos_24_28_semanas': partos_24_28_semanas,
        'partos_28_32_semanas': partos_28_32_semanas,
        'partos_32_37_semanas': partos_32_37_semanas,
        'lactancia_60min': lactancia_60min,
        'pueblos_originarios': pueblos_originarios,
        'migrantes': migrantes,
        'profilaxis_ocular_total': profilaxis_ocular_total,
        'profilaxis_ocular_pueblos': profilaxis_ocular_pueblos,
        'profilaxis_ocular_migrantes': profilaxis_ocular_migrantes,
        'rn_hepatitis_b_con_profilaxis': rn_hepatitis_b_con_profilaxis,
        'rn_hepatitis_b_sin_profilaxis': rn_hepatitis_b_sin_profilaxis,
    }
    return render(request, 'administrativo/reportes_generales.html', context)

def _obtener_datos_reporte(fecha_desde, fecha_hasta):
    """Función auxiliar para obtener todos los datos del reporte"""
    recien_nacidos = RecienNacido.objects.filter(
        fecha_nacimiento__range=[fecha_desde, fecha_hasta]
    )
    
    total_nacidos_vivos = recien_nacidos.count()
    
    # Peso al nacer
    peso_menor_500 = recien_nacidos.filter(peso__lt=500).count()
    peso_500_999 = recien_nacidos.filter(peso__gte=500, peso__lt=1000).count()
    peso_1000_1499 = recien_nacidos.filter(peso__gte=1000, peso__lt=1500).count()
    peso_1500_1999 = recien_nacidos.filter(peso__gte=1500, peso__lt=2000).count()
    peso_2000_2499 = recien_nacidos.filter(peso__gte=2000, peso__lt=2500).count()
    peso_2500_2999 = recien_nacidos.filter(peso__gte=2500, peso__lt=3000).count()
    peso_3000_3999 = recien_nacidos.filter(peso__gte=3000, peso__lt=4000).count()
    peso_4000_mas = recien_nacidos.filter(peso__gte=4000).count()
    
    anomalia_congenita = recien_nacidos.filter(malformacion_congenita=True).count()
    
    # Profilaxis
    profilaxis_hepatitis_b = recien_nacidos.filter(vacuna_hepatitis_b=True).count()
    profilaxis_ocular = recien_nacidos.filter(profilaxis_ocular=True).count()
    
    # Tipo de parto
    parto_vaginal = recien_nacidos.filter(tipo_parto='normal').count()
    parto_instrumental = recien_nacidos.filter(tipo_parto='instrumental').count()
    cesarea = recien_nacidos.filter(tipo_parto='cesarea').count()
    
    # APGAR
    apgar_1min_menor_igual_3 = recien_nacidos.filter(apgar_1min__lte=3).count()
    apgar_5min_menor_igual_5 = recien_nacidos.filter(apgar_5min__lte=5).count()
    reanimacion_basica_count = recien_nacidos.filter(reanimacion_basica=True).count()
    reanimacion_avanzada_count = recien_nacidos.filter(reanimacion_avanzada=True).count()
    
    # Características del parto
    total_partos = recien_nacidos.count()
    partos_vaginal = recien_nacidos.filter(tipo_parto='normal').count()
    partos_instrumental = recien_nacidos.filter(tipo_parto='instrumental').count()
    cesarea_electiva = recien_nacidos.filter(tipo_parto='cesarea').count()
    cesarea_urgencia = 0
    
    # Partos según edad de la madre
    pacientes_madres = Paciente.objects.filter(
        recien_nacidos__fecha_nacimiento__range=[fecha_desde, fecha_hasta]
    ).distinct()
    
    partos_menor_15 = pacientes_madres.filter(edad__lt=15).count()
    partos_15_19 = pacientes_madres.filter(edad__gte=15, edad__lt=20).count()
    partos_20_34 = pacientes_madres.filter(edad__gte=20, edad__lt=35).count()
    partos_35_mas = pacientes_madres.filter(edad__gte=35).count()
    
    # Partos prematuros
    partos_22_24_semanas = recien_nacidos.filter(
        edad_gestacional_semanas__gte=22, edad_gestacional_semanas__lt=24
    ).count()
    partos_24_28_semanas = recien_nacidos.filter(
        edad_gestacional_semanas__gte=24, edad_gestacional_semanas__lt=28
    ).count()
    partos_28_32_semanas = recien_nacidos.filter(
        edad_gestacional_semanas__gte=28, edad_gestacional_semanas__lt=32
    ).count()
    partos_32_37_semanas = recien_nacidos.filter(
        edad_gestacional_semanas__gte=32, edad_gestacional_semanas__lt=37
    ).count()
    
    lactancia_60min = recien_nacidos.filter(
        lactancia_antes_60min=True,
        peso__gte=2500
    ).count()
    
    # Pueblos originarios y migrantes
    pueblos_originarios = pacientes_madres.filter(
        pueblo_originario__isnull=False
    ).exclude(pueblo_originario='').count()
    migrantes = pacientes_madres.filter(es_migrante=True).count()
    
    # Profilaxis ocular
    profilaxis_ocular_total = recien_nacidos.filter(profilaxis_ocular=True).count()
    profilaxis_ocular_pueblos = recien_nacidos.filter(
        profilaxis_ocular=True,
        paciente_madre__pueblo_originario__isnull=False
    ).exclude(paciente_madre__pueblo_originario='').count()
    profilaxis_ocular_migrantes = recien_nacidos.filter(
        profilaxis_ocular=True,
        paciente_madre__es_migrante=True
    ).count()
    
    # Profilaxis Hepatitis B
    madres_hepatitis_b = pacientes_madres.filter(hepatitis_b_resultado__icontains='positivo')
    rn_hepatitis_b = recien_nacidos.filter(paciente_madre__in=madres_hepatitis_b)
    rn_hepatitis_b_con_profilaxis = rn_hepatitis_b.filter(
        vacuna_hepatitis_b=True,
        paciente_madre__hepatitis_b_profilaxis_completa=True
    ).count()
    rn_hepatitis_b_sin_profilaxis = rn_hepatitis_b.count() - rn_hepatitis_b_con_profilaxis
    
    return {
        'total_nacidos_vivos': total_nacidos_vivos,
        'peso_menor_500': peso_menor_500,
        'peso_500_999': peso_500_999,
        'peso_1000_1499': peso_1000_1499,
        'peso_1500_1999': peso_1500_1999,
        'peso_2000_2499': peso_2000_2499,
        'peso_2500_2999': peso_2500_2999,
        'peso_3000_3999': peso_3000_3999,
        'peso_4000_mas': peso_4000_mas,
        'anomalia_congenita': anomalia_congenita,
        'profilaxis_hepatitis_b': profilaxis_hepatitis_b,
        'profilaxis_ocular': profilaxis_ocular,
        'parto_vaginal': parto_vaginal,
        'parto_instrumental': parto_instrumental,
        'cesarea': cesarea,
        'apgar_1min_menor_igual_3': apgar_1min_menor_igual_3,
        'apgar_5min_menor_igual_5': apgar_5min_menor_igual_5,
        'reanimacion_basica_count': reanimacion_basica_count,
        'reanimacion_avanzada_count': reanimacion_avanzada_count,
        'total_partos': total_partos,
        'partos_vaginal': partos_vaginal,
        'partos_instrumental': partos_instrumental,
        'cesarea_electiva': cesarea_electiva,
        'cesarea_urgencia': cesarea_urgencia,
        'partos_menor_15': partos_menor_15,
        'partos_15_19': partos_15_19,
        'partos_20_34': partos_20_34,
        'partos_35_mas': partos_35_mas,
        'partos_22_24_semanas': partos_22_24_semanas,
        'partos_24_28_semanas': partos_24_28_semanas,
        'partos_28_32_semanas': partos_28_32_semanas,
        'partos_32_37_semanas': partos_32_37_semanas,
        'lactancia_60min': lactancia_60min,
        'pueblos_originarios': pueblos_originarios,
        'migrantes': migrantes,
        'profilaxis_ocular_total': profilaxis_ocular_total,
        'profilaxis_ocular_pueblos': profilaxis_ocular_pueblos,
        'profilaxis_ocular_migrantes': profilaxis_ocular_migrantes,
        'rn_hepatitis_b_con_profilaxis': rn_hepatitis_b_con_profilaxis,
        'rn_hepatitis_b_sin_profilaxis': rn_hepatitis_b_sin_profilaxis,
    }

@login_required
def exportar_reporte_pdf(request):
    """Exportar reporte a PDF"""
    if not REPORTLAB_AVAILABLE:
        from django.contrib import messages
        messages.error(request, 'La librería reportlab no está instalada. Ejecute: pip install reportlab')
        return redirect('reportes_generales')
    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')
    
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
        except:
            fecha_desde = timezone.now().replace(day=1).date()
    else:
        fecha_desde = timezone.now().replace(day=1).date()
    
    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
        except:
            fecha_hasta = timezone.now().date()
    else:
        fecha_hasta = timezone.now().date()
    
    datos = _obtener_datos_reporte(fecha_desde, fecha_hasta)
    
    # Crear el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    elements.append(Paragraph("REPORTES GENERALES DEL SISTEMA", title_style))
    elements.append(Paragraph("Hospital Clínico Herminda Martín — Chillán", styles['Normal']))
    elements.append(Paragraph(f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # SECCIÓN D.1
    elements.append(Paragraph("SECCIÓN D.1: INFORMACIÓN GENERAL DE RECIÉN NACIDOS VIVOS", styles['Heading2']))
    data_d1 = [
        ['TIPO', 'TOTAL', '<500', '500-999', '1.000-1.499', '1.500-1.999', '2.000-2.499', '2.500-2.999', '3.000-3.999', '4.000 y más', 'Anomalía Congénita'],
        ['NACIDOS VIVOS', str(datos['total_nacidos_vivos']), str(datos['peso_menor_500']), 
         str(datos['peso_500_999']), str(datos['peso_1000_1499']), str(datos['peso_1500_1999']),
         str(datos['peso_2000_2499']), str(datos['peso_2500_2999']), str(datos['peso_3000_3999']),
         str(datos['peso_4000_mas']), str(datos['anomalia_congenita'])]
    ]
    
    table_d1 = Table(data_d1)
    table_d1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(table_d1)
    elements.append(Spacer(1, 0.2*inch))
    
    # SECCIÓN D.2
    elements.append(Paragraph("SECCIÓN D.2: ATENCIÓN INMEDIATA DEL RECIÉN NACIDO", styles['Heading2']))
    data_d2 = [
        ['TIPO', 'Hepatitis B', 'Ocular', 'Parto Vaginal', 'Parto Instrumental', 'Cesárea', 
         '1 min ≤3', '5 min ≤5', 'Reanimación Básica', 'Reanimación Avanzada'],
        ['NACIDOS VIVOS', str(datos['profilaxis_hepatitis_b']), str(datos['profilaxis_ocular']),
         str(datos['parto_vaginal']), str(datos['parto_instrumental']), str(datos['cesarea']),
         str(datos['apgar_1min_menor_igual_3']), str(datos['apgar_5min_menor_igual_5']),
         str(datos['reanimacion_basica_count']), str(datos['reanimacion_avanzada_count'])]
    ]
    
    table_d2 = Table(data_d2)
    table_d2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(table_d2)
    elements.append(Spacer(1, 0.2*inch))
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_general_{fecha_desde}_{fecha_hasta}.pdf"'
    return response

@login_required
def exportar_reporte_excel(request):
    """Exportar reporte a Excel"""
    if not OPENPYXL_AVAILABLE:
        from django.contrib import messages
        messages.error(request, 'La librería openpyxl no está instalada. Ejecute: pip install openpyxl')
        return redirect('reportes_generales')
    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')
    
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
        except:
            fecha_desde = timezone.now().replace(day=1).date()
    else:
        fecha_desde = timezone.now().replace(day=1).date()
    
    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
        except:
            fecha_hasta = timezone.now().date()
    else:
        fecha_hasta = timezone.now().date()
    
    datos = _obtener_datos_reporte(fecha_desde, fecha_hasta)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte General"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:K1')
    ws['A1'] = "REPORTES GENERALES DEL SISTEMA"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:K2')
    ws['A2'] = "Hospital Clínico Herminda Martín — Chillán"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:K3')
    ws['A3'] = f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # SECCIÓN D.1
    row = 5
    ws.merge_cells(f'A{row}:K{row}')
    ws[f'A{row}'] = "SECCIÓN D.1: INFORMACIÓN GENERAL DE RECIÉN NACIDOS VIVOS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    headers_d1 = ['TIPO', 'TOTAL', '<500', '500-999', '1.000-1.499', '1.500-1.999', 
                  '2.000-2.499', '2.500-2.999', '3.000-3.999', '4.000 y más', 'Anomalía Congénita']
    for col, header in enumerate(headers_d1, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    data_d1 = ['NACIDOS VIVOS', datos['total_nacidos_vivos'], datos['peso_menor_500'],
               datos['peso_500_999'], datos['peso_1000_1499'], datos['peso_1500_1999'],
               datos['peso_2000_2499'], datos['peso_2500_2999'], datos['peso_3000_3999'],
               datos['peso_4000_mas'], datos['anomalia_congenita']]
    
    for col, value in enumerate(data_d1, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # SECCIÓN D.2
    row += 3
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = "SECCIÓN D.2: ATENCIÓN INMEDIATA DEL RECIÉN NACIDO"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    headers_d2 = ['TIPO', 'Hepatitis B', 'Ocular', 'Parto Vaginal', 'Parto Instrumental', 
                  'Cesárea', '1 min ≤3', '5 min ≤5', 'Reanimación Básica', 'Reanimación Avanzada']
    for col, header in enumerate(headers_d2, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    data_d2 = ['NACIDOS VIVOS', datos['profilaxis_hepatitis_b'], datos['profilaxis_ocular'],
               datos['parto_vaginal'], datos['parto_instrumental'], datos['cesarea'],
               datos['apgar_1min_menor_igual_3'], datos['apgar_5min_menor_igual_5'],
               datos['reanimacion_basica_count'], datos['reanimacion_avanzada_count']]
    
    for col, value in enumerate(data_d2, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_general_{fecha_desde}_{fecha_hasta}.xlsx"'
    return response

@login_required
def lista_altas_egresos(request):
    """Lista todas las altas y egresos"""
    altas = AltaEgreso.objects.all().select_related('paciente', 'admision', 'usuario_registro').order_by('-fecha_alta')
    
    # Filtros
    estado_filter = request.GET.get('estado', '')
    tipo_alta_filter = request.GET.get('tipo_alta', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if estado_filter:
        altas = altas.filter(estado_episodio=estado_filter)
    if tipo_alta_filter:
        altas = altas.filter(tipo_alta=tipo_alta_filter)
    if fecha_desde:
        altas = altas.filter(fecha_alta__date__gte=fecha_desde)
    if fecha_hasta:
        altas = altas.filter(fecha_alta__date__lte=fecha_hasta)
    
    context = {
        'altas': altas,
        'estado_filter': estado_filter,
        'tipo_alta_filter': tipo_alta_filter,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    return render(request, 'administrativo/lista_altas_egresos.html', context)

@login_required
def registrar_alta(request, admision_id=None):
    """Registrar alta de un paciente"""
    admision = None
    if admision_id:
        admision = get_object_or_404(Admision, id=admision_id)
    
    # Verificar si ya existe un alta para esta admisión
    if admision and hasattr(admision, 'alta_egreso'):
        messages.info(request, 'Esta admisión ya tiene un alta registrada.')
        return redirect('detalle_alta_egreso', id=admision.alta_egreso.id)
    
    if request.method == 'POST':
        form = AltaEgresoForm(request.POST)
        if form.is_valid():
            alta = form.save(commit=False)
            alta.usuario_registro = request.user
            if admision:
                alta.admision = admision
                alta.paciente = admision.paciente
                # Actualizar fecha de alta en la admisión
                admision.fecha_alta = timezone.now()
                admision.estado = 'completada'
                admision.save()
            alta.save()
            messages.success(request, f'Alta registrada exitosamente para {alta.paciente.nombre}.')
            return redirect('detalle_alta_egreso', id=alta.id)
    else:
        initial = {}
        if admision:
            initial['admision'] = admision
            initial['paciente'] = admision.paciente
        form = AltaEgresoForm(initial=initial)
    
    context = {
        'form': form,
        'admision': admision,
    }
    return render(request, 'administrativo/registrar_alta.html', context)

@login_required
def detalle_alta_egreso(request, id):
    """Detalle de una alta y egreso"""
    alta = get_object_or_404(AltaEgreso, id=id)
    
    context = {
        'alta': alta,
    }
    return render(request, 'administrativo/detalle_alta_egreso.html', context)

@login_required
def cerrar_episodio(request, id):
    """Cerrar un episodio de alta"""
    alta = get_object_or_404(AltaEgreso, id=id)
    
    if request.method == 'POST':
        alta.estado_episodio = 'cerrado'
        alta.save()
        messages.success(request, f'Episodio cerrado exitosamente para {alta.paciente.nombre}.')
        return redirect('detalle_alta_egreso', id=alta.id)
    
    return render(request, 'administrativo/cerrar_episodio.html', {'alta': alta})

@login_required
def entregar_documentos(request, id):
    """Registrar entrega de documentos a la familia"""
    alta = get_object_or_404(AltaEgreso, id=id)
    
    if request.method == 'POST':
        alta.documento_alta_entregado = request.POST.get('documento_alta_entregado') == 'on'
        alta.certificado_defuncion_entregado = request.POST.get('certificado_defuncion_entregado') == 'on'
        alta.informe_medico_entregado = request.POST.get('informe_medico_entregado') == 'on'
        alta.recetas_entregadas = request.POST.get('recetas_entregadas') == 'on'
        alta.otros_documentos = request.POST.get('otros_documentos', '')
        alta.persona_que_recibio = request.POST.get('persona_que_recibio', '')
        
        fecha_entrega = request.POST.get('fecha_entrega_documentos', '')
        if fecha_entrega:
            try:
                alta.fecha_entrega_documentos = datetime.strptime(fecha_entrega, '%Y-%m-%dT%H:%M')
            except:
                alta.fecha_entrega_documentos = timezone.now()
        else:
            alta.fecha_entrega_documentos = timezone.now()
        
        alta.save()
        messages.success(request, 'Entrega de documentos registrada exitosamente.')
        return redirect('detalle_alta_egreso', id=alta.id)
    
    return render(request, 'administrativo/entregar_documentos.html', {'alta': alta})

@login_required
def tramites_administrativos(request, id):
    """Registrar trámites administrativos del egreso"""
    alta = get_object_or_404(AltaEgreso, id=id)
    
    if request.method == 'POST':
        alta.pagos_completados = request.POST.get('pagos_completados') == 'on'
        alta.facturacion_cerrada = request.POST.get('facturacion_cerrada') == 'on'
        alta.habitacion_liberada = request.POST.get('habitacion_liberada') == 'on'
        alta.expediente_clinico_archivado = request.POST.get('expediente_clinico_archivado') == 'on'
        alta.notificaciones_enviadas = request.POST.get('notificaciones_enviadas') == 'on'
        alta.tramites_adicionales = request.POST.get('tramites_adicionales', '')
        alta.observaciones = request.POST.get('observaciones', '')
        alta.save()
        messages.success(request, 'Trámites administrativos registrados exitosamente.')
        return redirect('detalle_alta_egreso', id=alta.id)
    
    return render(request, 'administrativo/tramites_administrativos.html', {'alta': alta})
